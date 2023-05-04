""" Create handy calendars to print from Excel

Michel van Wouw: I used to make these by hand every year, print them out and stick them underneath my monitor at work.
                 They give an easy way to mark dates and see day types and weeknumbers at a glance. Much
                 over-engineered for what it does, but it was a fun exercise.

Usage:
    XLCalendar [options]

General Options:
    -h, --help          Display help.
    -v, --version       Display version.
    -s <M> <YYYY>       Start calendar with month <M> of year <YYYY>. (Default: month 1 of current year)
    -e <M> <YYYY>       End calendar with month <M> of year <YYYY>. (Default: month 1 of next year)
    -o <output_file>    Use <output_file> as filename. (Default: 'Calendar.xlsx')
    -wr <%>             Resize column widths to <%> percent.
    -hr <%>             Resize row heights to <%> percent.
    -fnl                Force day and month names to NL. (Default: OS locale)
    -mnl                Mark NL general holidays.
"""


import datetime
import re
from math import floor
from sys import argv, exit
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


# Default settings
opt = {
    "Y_S": int(datetime.date.today().strftime("%Y")),
    "M_S": 1,
    "Y_E": int(datetime.date.today().strftime("%Y")) + 1,
    "M_E": 1,
    "COLUMN_WIDTH": 3.5,
    "ROW_HEIGHT": 12.7,
    "FORCE_NL": False,
    "OUTPUT_FILE": f"Calendar.xlsx",
    "OUTPUT_FILE_SET": False,
    "VERSION": 1.11,
    "NL_HOLIDAYS": False,
    "HELP_TEXT": """
Tiny Calendar - Create handy calendars to print from Excel

Usage:
    XLCalendar [options]

General Options:
    -h, --help          Display help.
    -v, --version       Display version.
    -s <M> <YYYY>       Start calendar with month <M> of year <YYYY>. (Default: month 1 of current year)
    -e <M> <YYYY>       End calendar with month <M> of year <YYYY>. (Default: month 1 of next year)
    -o <output_file>    Use <output_file> as filename. (Default: 'Calendar.xlsx')
    -wr <%>             Resize column widths to <%> percent.
    -hr <%>             Resize row heights to <%> percent.
    -fnl                Force day and month names to NL. (Default: OS locale)
    -mnl                Mark NL general holidays.
"""
}

# Define styling templates
font_standard = Font(name="Arial", size=10, bold=False)
font_bold = Font(name="Arial", size=10, bold=True)

h_align_center = Alignment(horizontal="center", vertical="center")
h_align_right = Alignment(horizontal="right", vertical="center")

fill_white = PatternFill("solid", fgColor="FFFFFF")
fill_lightgrey = PatternFill("solid", fgColor="F2F2F2")
fill_mediumgrey = PatternFill("solid", fgColor="D9D9D9")

thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")

border_LrTb = Border(left=medium, right=thin, top=medium, bottom=thin)
border_Lr = Border(left=medium, right=thin)
border_LrtB = Border(left=medium, right=thin, top=thin, bottom=medium)
border_lrTb = Border(left=thin, right=thin, top=medium, bottom=thin)
border_lRTb = Border(left=thin, right=medium, top=medium, bottom=thin)
border_lt = Border(left=thin, top=thin)
border_l = Border(left=thin)
border_R = Border(right=medium)
border_lrtB = Border(left=thin, right=thin, top=thin, bottom=medium)
border_lRtB = Border(left=thin, right=medium, top=thin, bottom=medium)
border_lRt = Border(left=thin, right=medium, top=thin)
border_lR = Border(left=thin, right=medium)


def main() -> None:
    # Process command line arguments
    if len(argv) == 1:
        # No command line arguments -> use default settings
        print(f"Generating Tiny Calendar with default settings. Use 'XLCalendar -h' for options.\n")
        create_calendar_file()
    else:
        # One or more command line arguments
        # could have used .pop(0) i guess :/
        cl_args = argv[1:]
        cl_args = cl_args[::-1]
        while cl_args:
            # Go through provided arguments and validate
            arg = cl_args.pop()

            # Option: help
            if arg == "-h" or arg == "--help":
                hae()

            # Option: version
            if arg == "-v" or arg == "--version":
                print(f"\nXLCalendar version {opt['VERSION']}")
                exit()

            # Option: -s <M> <YYYY>
            elif arg == "-s":
                try:
                    sm = int(cl_args.pop())
                    if 1 <= sm <= 12:
                        opt['M_S'] = sm
                    else:
                        print(f"\nERROR: Option '-s' positional argument <M> should be a number from 1 to 12.")
                        hae()
                    sy = int(cl_args.pop())
                    if sy > 0:
                        opt['Y_S'] = sy
                    else:
                        print(f"\nERROR: Option '-s' positional argument <YYYY> should be a number > 0.")
                        hae()
                    if not opt['OUTPUT_FILE_SET']:
                        opt['OUTPUT_FILE'] = f"Calendar {opt['M_S']}-{opt['Y_S']} to {opt['M_E']}-{opt['Y_E']}.xlsx"
                except IndexError:
                    print(f"\nERROR: Option '-s' requires positional arguments <M> and <YYYY>.")
                    hae()
                except ValueError:
                    print(f"\nERROR: Option '-s' positional arguments <M> (1-12) and <YYYY> (1-inf) should be numbers.")
                    hae()

            # Option: -e <M> <YYYY>
            elif arg == "-e":
                try:
                    em = int(cl_args.pop())
                    if 1 <= em <= 12:
                        opt['M_E'] = em
                    else:
                        print(f"\nERROR: Option '-e' positional argument <M> should be a number from 1 to 12.")
                        hae()
                    ey = int(cl_args.pop())
                    if ey > 0:
                        opt['Y_E'] = ey
                    else:
                        print(f"\nERROR: Option '-e' positional argument <YYYY> should be a number > 0.")
                        hae()
                    if not opt['OUTPUT_FILE_SET']:
                        opt['OUTPUT_FILE'] = f"Calendar {opt['M_S']}-{opt['Y_S']} to {opt['M_E']}-{opt['Y_E']}.xlsx"
                except IndexError:
                    print(f"\nERROR: Option '-e' requires positional arguments <M> and <YYYY>.")
                    hae()
                except ValueError:
                    print(f"\nERROR: Option '-e' positional arguments <M> (1-12) and <YYYY> (1-inf) should be numbers.")
                    hae()

            # Option: -o <output_file>
            elif arg == "-o":
                try:
                    file_name = cl_args.pop()
                    if not re.fullmatch(r"^[0-9a-zA-Z_\-][0-9a-zA-Z_\-. ]*$", file_name):
                        print(f"\nERROR: Provided filename is not a valid Windows filename.")
                        hae()
                    opt['OUTPUT_FILE'] = file_name if re.fullmatch(r"^.*\.xlsx$", file_name) else file_name + ".xlsx"
                    opt['OUTPUT_FILE_SET'] = True
                except IndexError:
                    print(f"\nERROR: Option '-o' requires positional argument <output_file>.")
                    hae()

            # Option: -wr <%>
            elif arg == "-wr":
                try:
                    wr = cl_args.pop()
                    if int(wr) > 0:
                        opt['COLUMN_WIDTH'] = float(opt['COLUMN_WIDTH'] * (int(wr) / 100))
                    else:
                        print(f"\nERROR: Option '-wr' positional argument <%> should be a number > 0.")
                        hae()
                except IndexError:
                    print(f"\nERROR: Option '-wr' requires positional argument <%>.")
                    hae()
                except ValueError:
                    print(f"\nERROR: Option '-wr' positional argument <%> should be a number > 0.")
                    hae()

            # Option: -hr <%>
            elif arg == "-hr":
                try:
                    hr = cl_args.pop()
                    if int(hr) > 0:
                        opt['ROW_HEIGHT'] = float(opt['ROW_HEIGHT'] * (int(hr) / 100))
                    else:
                        print(f"\nERROR: Option '-hr' positional argument <%> should be a number > 0.")
                        hae()
                except IndexError:
                    print(f"\nERROR: Option '-hr' requires positional argument <%>.")
                    hae()
                except ValueError:
                    print(f"\nERROR: Option '-hr' positional argument <%> should be a number > 0.")
                    hae()

            # Option: -fnl
            elif arg == "-fnl":
                opt['FORCE_NL'] = True

            # Option: -mnl
            elif arg == "-mnl":
                opt['NL_HOLIDAYS'] = True

            # Unsupported argument provided
            else:
                print(f"\nERROR: No such option: {arg}")
                hae()

        # Check whether end date is at least the same month or later
        if opt['Y_E'] < opt['Y_S']:
            print(f"\nERROR: End date is earlyer than start date.")
            hae()
        elif opt['Y_E'] == opt['Y_S']:
            if opt['M_E'] < opt['M_S']:
                print(f"\nERROR: End date is earlyer than start date.")
                hae()
        # Check maximum span of full calendar
        elif opt['Y_E'] - opt['Y_S'] > 100:
            print(f"\nERROR: Calendar cannot span more than 100 years.")
            hae()

        # All options parsed and validated -> create the file
        print(f"\nCreating calendar running from {opt['M_S']}-{opt['Y_S']} to {opt['M_E']}-{opt['Y_E']}.")
        create_calendar_file()


def hae() -> None:
    # Display help text and exit
    print(opt['HELP_TEXT'])
    exit()


def get_week_info(date_in: datetime.date) -> tuple[int, int]:
    # Returns the weeknumber and day-of-the-week for the given day
    _, week_number, week_day = date_in.isocalendar()
    return week_number, week_day


def get_easter_date(year: int) -> datetime.date:
    # Returns a date object of the first easter-day of the argumented year
    # Thank you rahulhegde97 for showing me this algorithm
    # https://www.geeksforgeeks.org/how-to-calculate-the-easter-date-for-a-given-year-using-gauss-algorithm/

    # Calculate the location of the year in the Metonic cycle
    metonic = year % 19
    # Find the number of leap days according to Julian’s calendar
    leapdays = year % 4
    # Take into account that the non-leap year is one day longer than 52 weeks
    remdays = year % 7
    # m depends on the century of year. For 19th century, m = 23. For the 21st century, m = 24 and so on
    p = floor(year / 100)
    q = floor((13 + 8 * p) / 25)
    m = (15 - q + p - p // 4) % 30
    # The difference between the number of leap days between the Julian and the Gregorian calendar is given by:
    n = (4 + p - p // 4) % 7
    # The number of days to be added to March 21 to find the date of the Paschal Full Moon is given by:
    d = (19 * metonic + m) % 30
    # The number of days from the Paschal full moon to the next Sunday is given by:
    e = (2 * leapdays + 4 * remdays + 6 * d + n) % 7
    # Using d and e, the date of Easter Sunday is going to be March (22 + d + e). If this number comes out to be
    # greater than 31, then we move to April.
    days = (22 + d + e)
    # Now the lunar month is not exactly 30 days but a little less than 30 days. To nullify this inconsistency the
    # following cases are applied:
    if d == 29 and e == 6:
        return datetime.date(year, 4, 19)
    elif d == 28 and e == 6:
        return datetime.date(year, 4, 18)
    else:
        # If days > 31, move to April
        if days > 31:
            return datetime.date(year, 4, (days - 31))
        else:
            # Otherwise, stay in March
            return datetime.date(year, 3, days)


def get_holidays(year_start: int, year_end: int) -> list[datetime.date]:
    # Returns a list with date objects of all NL holidays in the given timespan
    holidays = []
    for year in range(year_start, year_end + 1):
        # New Year's Day
        holidays.append(datetime.date(year, 1, 1))
        # 1st and 2nd Easter days
        easter = get_easter_date(year)
        holidays.extend([easter, easter + datetime.timedelta(days=1)])
        # Good Friday
        holidays.append(easter - datetime.timedelta(days=2))
        # Ascension Day
        holidays.append(easter + datetime.timedelta(days=39))
        # NL Kings Day
        holidays.append(datetime.date(year, 4, 27))
        # 1st and 2nd Pentecost Day
        holidays.append(easter + datetime.timedelta(days=49))
        holidays.append(easter + datetime.timedelta(days=50))
        # 1st and 2nd Christmas Day
        holidays.append(datetime.date(year, 12, 25))
        holidays.append(datetime.date(year, 12, 26))
    return holidays


def create_calendar_file() -> None:
    # Get localized day and month names
    if opt['FORCE_NL']:
        weekday_dict = {1: "Ma", 2: "Di", 3: "Wo", 4: "Do", 5: "Vr", 6: "Za", 7: "Zo", 8: "Week"}
        month_dict = {1: "Januari", 2: "Februari", 3: "Maart", 4: "April", 5: "Mei", 6: "Juni", 7: "Juli",
                      8: "Augustus", 9: "September", 10: "Oktober", 11: "November", 12: "December"}
    else:
        weekday_dict = {}
        i = 1
        for day in range(24, 31):
            d = datetime.date(2023, 4, day)
            weekday_dict[i] = d.strftime("%a")
            i += 1
        # weekday_dict[8] = "Week"
        month_dict = {}
        for month in range(1, 13):
            d = datetime.date(2023, month, 1)
            month_dict[month] = d.strftime("%B")

    # Gather the list of holidays for marking
    if opt['NL_HOLIDAYS']:
        holidays = get_holidays(opt['Y_S'], opt['Y_E'])

    # Get a list of days for the requested timespan
    # per day: [datetime.date, weeknumber, weekday, is-holiday]
    day_list = []
    td = datetime.timedelta(days=1)
    current_day = datetime.date(opt['Y_S'], opt['M_S'], 1)
    if opt['M_E'] == 12:
        last_day = datetime.date(opt['Y_E'] + 1, 1, 1) - td
    else:
        last_day = datetime.date(opt['Y_E'], opt['M_E'] + 1, 1) - td

    if opt['NL_HOLIDAYS']:
        while current_day <= last_day:
            if current_day in holidays:
                day_list.append([current_day, *get_week_info(current_day), 1])
            else:
                day_list.append([current_day, *get_week_info(current_day), 0])
            current_day += td
    else:
        while current_day <= last_day:
            day_list.append([current_day, *get_week_info(current_day), 0])
            current_day += td

    # Add padding days to get full start and end weeks
    # Front padding
    while day_list[0][2] != 1:
        current_day = day_list[0][0] - datetime.timedelta(days=1)
        if opt['NL_HOLIDAYS']:
            if current_day in holidays:
                day_list.insert(0, [current_day, *get_week_info(current_day), 1])
            else:
                day_list.insert(0, [current_day, *get_week_info(current_day), 0])
        else:
            day_list.insert(0, [current_day, *get_week_info(current_day), 0])
    # Back padding
    while day_list[-1][2] != 7:
        current_day = day_list[-1][0] + datetime.timedelta(days=1)
        if opt['NL_HOLIDAYS']:
            if current_day in holidays:
                day_list.append([current_day, *get_week_info(current_day), 1])
            else:
                day_list.append([current_day, *get_week_info(current_day), 0])
        else:
            day_list.append([current_day, *get_week_info(current_day), 0])

    # Create an openpyxl worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Calendar {opt['M_S']}-{opt['Y_S']} to {opt['M_E']}-{opt['Y_E']}"

    # Write row header values and styles
    # First year
    c = ws.cell(row=1, column=1)
    c.value = opt['Y_S']
    c.font = font_bold
    c.alignment = h_align_center
    c.fill = fill_white
    c.border = border_LrTb
    # Short day names
    for i in range(1, 8):
        c = ws.cell(row=i + 1, column=1)
        c.value = weekday_dict[i]
        c.font = font_standard
        c.alignment = h_align_right
        c.fill = fill_white
        c.border = border_Lr
    # 'week' cell
    c = ws.cell(row=9, column=1)
    # c.value = weekday_dict[8]
    c.font = font_standard
    c.alignment = h_align_right
    c.fill = fill_white
    c.border = border_LrtB
    # Merge the row header cells
    for i in range(1, 10):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

    # Get some limits for further filling and styling the sheet
    len_day_list = len(day_list)
    last_column = int(len_day_list / 7 + 2)

    # Write all month headers + day- and weeknumbers from the list of days to the worksheet
    day_index = 0
    for col in range(3, last_column + 1):
        if day_index == len_day_list:
            break
        # Write month column header
        c = ws.cell(row=1, column=col)
        c.value = month_dict[day_list[day_index][0].month]
        c.font = font_bold
        c.alignment = h_align_center
        c.fill = fill_white

        # Write day- and weeknumber
        for row in range(1, 8):
            if day_list[day_index][2] == row:
                # days
                c = ws.cell(row=row + 1, column=col)
                c.value = day_list[day_index][0].day
                c.font = font_standard
                c.alignment = h_align_center
                # Fill mediumgrey for holidays, lightgrey for weekends, white for normal days
                if opt['NL_HOLIDAYS']:
                    if day_list[day_index][0] in holidays:
                        c.fill = fill_mediumgrey
                    elif day_list[day_index][2] == 6 or day_list[day_index][2] == 7:
                        c.fill = fill_lightgrey
                    else:
                        c.fill = fill_white
                else:
                    if day_list[day_index][2] == 6 or day_list[day_index][2] == 7:
                        c.fill = fill_lightgrey
                    else:
                        c.fill = fill_white

                # weeks
                c = ws.cell(row=9, column=col)
                c.value = day_list[day_index][1]
                c.font = font_standard
                c.alignment = h_align_center
                c.fill = fill_white

                day_index += 1
                if day_index == len_day_list:
                    break
            else:
                c = ws.cell(row=row + 1, column=col)
                c.font = font_standard
                c.alignment = h_align_center
                c.fill = fill_white

    # Merge header cells of corresponding months and set border styles
    prev_cell_content = ws.cell(row=1, column=3).value
    merge_range_start = 3
    merge_range_end = 3
    ws.cell(row=1, column=3).border = border_lrTb
    for col in range(4, last_column + 2):
        if ws.cell(row=1, column=col).value == prev_cell_content:
            merge_range_end = col
        else:
            if col == last_column + 1:
                ws.cell(row=1, column=merge_range_start).border = border_lRTb
                ws.merge_cells(start_row=1, start_column=merge_range_start, end_row=1, end_column=merge_range_end)
            else:
                ws.cell(row=1, column=col).border = border_lrTb
                ws.merge_cells(start_row=1, start_column=merge_range_start, end_row=1, end_column=merge_range_end)
                prev_cell_content = ws.cell(row=1, column=col).value
                if merge_range_start == 3 and merge_range_end == 3:
                    ws.cell(row=1, column=3).value = None
                merge_range_start = col

    # Apply column width and row height
    for i in range(1, last_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = opt['COLUMN_WIDTH']
    for i in range(1, 10):
        ws.row_dimensions[i].height = opt['ROW_HEIGHT']

    # Apply borders to daynumber grid
    for col in range(3, last_column):
        for row in range(2, 9):
            c = ws.cell(row=row, column=col)
            if c.value == 1:
                c.border = border_lt
            elif 2 <= ws.cell(row=row, column=col).value <= 7:
                c.border = border_l
    for row in range(2, 9):
        c = ws.cell(row=row, column=last_column)
        if c.value == 1:
            c.border = border_lRt
        elif 2 <= c.value <= 7:
            c.border = border_lR
        else:
            c.border = border_R

    # Apply borders for weeknumbers row
    for col in range(3, last_column):
        ws.cell(row=9, column=col).border = border_lrtB
    ws.cell(row=9, column=last_column).border = border_lRtB

    # Apply page setup and set printing options
    ws.print_area = f"A1:{get_column_letter(last_column)}9"
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True

    # Save file to disk
    wb.save(opt['OUTPUT_FILE'])
    print(f"\nFile saved as: '{opt['OUTPUT_FILE']}'")


if __name__ == "__main__":
    main()
